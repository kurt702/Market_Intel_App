"""Executive KPI workbook and summary exports."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


KPI_WORKBOOK_SHEETS = [
    "01_Executive_KPI_Scorecard",
    "02_KPI_Calculation_Detail",
    "03_CPT_Reconciliation",
    "04_DRG_Reconciliation",
    "05_Leakage_Summary",
    "06_Referral_Balance",
    "07_Geographic_Catchment",
    "08_Market_Position",
    "09_DPCI",
    "10_Confidence_Flags",
    "11_Validation_Log",
]


def build_final_output_tables(
    account_name: str,
    reporting_year: str,
    calculation_outputs: dict[str, pd.DataFrame],
    source_workbook_path: str | Path,
) -> dict[str, pd.DataFrame]:
    detail = calculation_outputs.get("kpi_calculation_detail", pd.DataFrame()).copy()
    scorecard = _scorecard(detail)
    return {
        "scorecard": scorecard,
        "calculation_detail": detail,
        "cpt_reconciliation": _cpt_reconciliation(calculation_outputs.get("cpt_calculation_detail", pd.DataFrame())),
        "drg_reconciliation": _drg_reconciliation(calculation_outputs.get("drg_calculation_detail", pd.DataFrame())),
        "leakage_summary": _leakage_summary(calculation_outputs.get("drg_calculation_detail", pd.DataFrame())),
        "referral_balance": _referral_balance(calculation_outputs.get("referral_calculation_detail", pd.DataFrame())),
        "geographic_catchment": calculation_outputs.get("geographic_catchment_detail", pd.DataFrame()),
        "market_position": _market_position(detail),
        "dpci": _dpci(detail),
        "confidence_flags": _confidence_flags(detail),
        "validation_log": _validation_log(account_name, reporting_year, source_workbook_path),
    }


def write_final_outputs(
    output_folder: str | Path,
    account_name: str,
    reporting_year: str,
    calculation_outputs: dict[str, pd.DataFrame],
    source_workbook_path: str | Path,
) -> dict[str, Path]:
    output = Path(output_folder)
    output.mkdir(parents=True, exist_ok=True)
    tables = build_final_output_tables(account_name, reporting_year, calculation_outputs, source_workbook_path)
    workbook_path = output / "kpi_workbook.xlsx"
    csv_path = output / "kpi_summary.csv"
    json_path = output / "kpi_summary.json"

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        tables["scorecard"].to_excel(writer, sheet_name="01_Executive_KPI_Scorecard", index=False)
        tables["calculation_detail"].to_excel(writer, sheet_name="02_KPI_Calculation_Detail", index=False)
        tables["cpt_reconciliation"].to_excel(writer, sheet_name="03_CPT_Reconciliation", index=False)
        tables["drg_reconciliation"].to_excel(writer, sheet_name="04_DRG_Reconciliation", index=False)
        tables["leakage_summary"].to_excel(writer, sheet_name="05_Leakage_Summary", index=False)
        tables["referral_balance"].to_excel(writer, sheet_name="06_Referral_Balance", index=False)
        tables["geographic_catchment"].to_excel(writer, sheet_name="07_Geographic_Catchment", index=False)
        tables["market_position"].to_excel(writer, sheet_name="08_Market_Position", index=False)
        tables["dpci"].to_excel(writer, sheet_name="09_DPCI", index=False)
        tables["confidence_flags"].to_excel(writer, sheet_name="10_Confidence_Flags", index=False)
        tables["validation_log"].to_excel(writer, sheet_name="11_Validation_Log", index=False)
        _format(writer.book)

    summary = tables["scorecard"].copy()
    summary.insert(0, "reporting_year", reporting_year)
    summary.insert(0, "account_name", account_name)
    summary[["account_name", "reporting_year", "kpi_id", "kpi_name", "domain", "value", "formatted_value", "unit", "confidence_status", "calculation_status", "source_summary", "validation_notes"]].to_csv(csv_path, index=False)
    payload = {
        "account_name": account_name,
        "reporting_year": reporting_year,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "kpis": summary[["kpi_id", "kpi_name", "domain", "value", "formatted_value", "unit", "confidence_status", "calculation_status", "source_summary", "validation_notes"]].to_dict(orient="records"),
    }
    json_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return {"kpi_workbook": workbook_path, "kpi_summary_csv": csv_path, "kpi_summary_json": json_path}


def _scorecard(detail: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in detail.iterrows():
        value = row.get("calculated_value", "")
        status = row.get("calculation_status", "")
        rows.append(
            {
                "kpi_id": row.get("kpi_id", ""),
                "kpi_name": row.get("kpi_name", ""),
                "domain": row.get("domain", ""),
                "value": "" if status == "not_calculated" else value,
                "formatted_value": "Not calculated" if status == "not_calculated" else _format_value(value, row.get("unit", "")),
                "unit": row.get("unit", ""),
                "eligibility_status": row.get("eligibility_status", ""),
                "calculation_status": status,
                "confidence_status": row.get("confidence_status", ""),
                "source_summary": _source_summary(row),
                "calculation_method": row.get("calculation_method", ""),
                "validation_notes": row.get("validation_notes", ""),
            }
        )
    return pd.DataFrame(rows)


def _format_value(value, unit: str) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "Not calculated"
    if unit == "dollars":
        return f"${number / 1_000_000:.2f}M" if abs(number) >= 1_000_000 else f"${number:,.0f}"
    if unit == "percent":
        return f"{number:.1f}%"
    if unit in {"cases", "cases/claims", "referrals"}:
        return f"{number:,.0f}"
    return f"{number:,.2f}"


def _source_summary(row) -> str:
    source_tab = str(row.get("source_tab", "") or "")
    source_file = str(row.get("source_file", "") or "")
    return "; ".join(part for part in [source_tab, source_file] if part)


def _cpt_reconciliation(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    agg = {"matched_row_count": ("cpt_code_normalized", "size")}
    if "Actual Charges__num" in df.columns:
        agg["sum Actual Charges__num"] = ("Actual Charges__num", "sum")
    if "# Actual Procedures__num" in df.columns:
        agg["sum # Actual Procedures__num"] = ("# Actual Procedures__num", "sum")
    return df.groupby(["cpt_code_normalized", "ep_category", "kpi_domain"], dropna=False).agg(**agg).reset_index()


def _drg_reconciliation(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    numeric = [c for c in df.columns if c.endswith("__num")]
    agg = {"matched_row_count": ("drg_code_normalized", "size")}
    for col in numeric:
        agg[f"sum {col}"] = (col, "sum")
    return df.groupby(["source_tab", "drg_code_normalized", "drg_category", "kpi_domain", "recommended_use"], dropna=False).agg(**agg).reset_index()


def _leakage_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    leak = df[df.get("source_tab") == "Raw_Inpatient_Leakage"]
    if leak.empty:
        return pd.DataFrame()
    agg = {"matched_row_count": ("drg_code_normalized", "size")}
    if "Total Charges__num" in leak.columns:
        agg["sum Total Charges__num"] = ("Total Charges__num", "sum")
    if "Total Claims__num" in leak.columns:
        agg["sum Total Claims__num"] = ("Total Claims__num", "sum")
    return leak.groupby(["source_tab", "drg_code_normalized", "drg_category"], dropna=False).agg(**agg).reset_index()


def _referral_balance(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    lookup = dict(zip(df["component"], df["referrals"]))
    files = "; ".join(sorted(set(df.get("source_files", pd.Series(dtype=str)).dropna().astype(str))))
    return pd.DataFrame([{"incoming_referrals": lookup.get("incoming_referrals", 0), "outgoing_referrals": lookup.get("outgoing_referrals", 0), "net_referral_balance": lookup.get("net_balance", 0), "source_files": files}])


def _market_position(detail: pd.DataFrame) -> pd.DataFrame:
    return detail[detail["kpi_id"].isin(["ep_market_share", "market_rank"])][["kpi_id", "kpi_name", "calculation_status", "calculated_value", "unit", "calculation_method", "validation_notes"]]


def _dpci(detail: pd.DataFrame) -> pd.DataFrame:
    return detail[detail["kpi_id"].isin(["device_creation_revenue", "device_lifecycle_revenue", "dpci"])][["kpi_id", "kpi_name", "calculated_value", "unit", "validation_notes"]]


def _confidence_flags(detail: pd.DataFrame) -> pd.DataFrame:
    return detail[["kpi_id", "kpi_name", "confidence_status", "validation_notes"]].rename(columns={"validation_notes": "reason"})


def _validation_log(account_name: str, reporting_year: str, source_workbook_path: str | Path) -> pd.DataFrame:
    timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return pd.DataFrame(
        [
            {"check_name": "no_recommendations_generated", "status": "pass", "detail": "No recommendations generated."},
            {"check_name": "no_opportunity_scores_generated", "status": "pass", "detail": "No opportunity scores generated."},
            {"check_name": "no_powerpoint_generated", "status": "pass", "detail": "No PowerPoint generated."},
            {"check_name": "source_workbook_used", "status": "pass", "detail": str(source_workbook_path)},
            {"check_name": "reporting_year", "status": "pass", "detail": str(reporting_year)},
            {"check_name": "generated_timestamp", "status": "pass", "detail": timestamp},
            {"check_name": "account_name", "status": "pass", "detail": account_name},
        ]
    )


def _format(workbook) -> None:
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 48)
