"""Auditor Workbook generation.

This module only writes the source fact book. It does not calculate KPIs or
create KPI summaries, presentations, recommendations, or opportunity scores.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .taxonomy import select_cpt_taxonomy_output_columns, select_drg_taxonomy_output_columns


RAW_TAB_SHEET_ORDER = [
    ("05_Raw_MedicalClaims_CPT", "Raw_MedicalClaims_CPT"),
    ("06_Raw_Outpatient_CPT", "Raw_Outpatient_CPT"),
    ("07_Raw_MedicalClaims_DRG", "Raw_MedicalClaims_DRG"),
    ("08_Raw_Inpatient_DRG", "Raw_Inpatient_DRG"),
    ("09_Raw_Inpatient_Leakage", "Raw_Inpatient_Leakage"),
    ("10_Raw_Outpatient_Leakage", "Raw_Outpatient_Leakage"),
    ("11_Raw_Inpatient_MarketShare", "Raw_Inpatient_MarketShare"),
    ("12_Raw_ZIP_Origination", "Raw_ZIP_Origination"),
    ("13_Raw_Referrals_From", "Raw_Referrals_From"),
    ("14_Raw_Referrals_To", "Raw_Referrals_To"),
    ("15_Raw_Unknown", "Raw_Unknown"),
]

AUDITOR_SHEET_NAMES = [
    "01_Read_Me",
    "02_Source_File_Inventory",
    "03_Raw_Tab_Summary",
    "04_Data_Quality_Log",
    *[sheet_name for sheet_name, _ in RAW_TAB_SHEET_ORDER],
    "16_CPT_Taxonomy_Matches",
    "17_DRG_Taxonomy_Matches",
    "19_KPI_Eligibility",
    "20_KPI_Calculation_Detail",
    "21_CPT_Calculation_Detail",
    "22_DRG_Calculation_Detail",
    "23_Referral_Calculation_Detail",
    "24_Geographic_Catchment_Detail",
]


def write_auditor_workbook(
    output_path: str | Path,
    account_name: str,
    source_file_inventory: pd.DataFrame,
    raw_tab_dataframes: dict[str, pd.DataFrame],
    raw_tab_summary: pd.DataFrame,
    cpt_taxonomy_matches: pd.DataFrame | None = None,
    drg_taxonomy_matches: pd.DataFrame | None = None,
    taxonomy_match_summary: pd.DataFrame | None = None,
    kpi_eligibility: pd.DataFrame | None = None,
    calculation_outputs: dict[str, pd.DataFrame] | None = None,
) -> Path:
    """Write the Auditor Workbook / source fact book."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    generation_timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _build_read_me(account_name, generation_timestamp).to_excel(writer, sheet_name="01_Read_Me", index=False, header=False)
        source_file_inventory.to_excel(writer, sheet_name="02_Source_File_Inventory", index=False)
        raw_tab_summary.to_excel(writer, sheet_name="03_Raw_Tab_Summary", index=False)
        _build_data_quality_log(
            source_file_inventory,
            raw_tab_dataframes,
            cpt_taxonomy_matches,
            drg_taxonomy_matches,
            taxonomy_match_summary,
            kpi_eligibility,
        ).to_excel(writer, sheet_name="04_Data_Quality_Log", index=False)

        for sheet_name, raw_tab_name in RAW_TAB_SHEET_ORDER:
            raw_df = raw_tab_dataframes.get(raw_tab_name, pd.DataFrame())
            if raw_df.empty:
                pd.DataFrame({"note": ["No source data loaded for this tab."]}).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                raw_df.to_excel(writer, sheet_name=sheet_name, index=False)

        cpt_output = select_cpt_taxonomy_output_columns(cpt_taxonomy_matches if cpt_taxonomy_matches is not None else pd.DataFrame())
        drg_output = select_drg_taxonomy_output_columns(drg_taxonomy_matches if drg_taxonomy_matches is not None else pd.DataFrame())
        if cpt_output.empty:
            pd.DataFrame({"note": ["No CPT taxonomy match data generated."]}).to_excel(writer, sheet_name="16_CPT_Taxonomy_Matches", index=False)
        else:
            cpt_output.to_excel(writer, sheet_name="16_CPT_Taxonomy_Matches", index=False)
        if drg_output.empty:
            pd.DataFrame({"note": ["No DRG taxonomy match data generated."]}).to_excel(writer, sheet_name="17_DRG_Taxonomy_Matches", index=False)
        else:
            drg_output.to_excel(writer, sheet_name="17_DRG_Taxonomy_Matches", index=False)
        if kpi_eligibility is None or kpi_eligibility.empty:
            pd.DataFrame({"note": ["No KPI eligibility data generated."]}).to_excel(writer, sheet_name="19_KPI_Eligibility", index=False)
        else:
            kpi_eligibility.to_excel(writer, sheet_name="19_KPI_Eligibility", index=False)
        _write_calculation_sheet(writer, calculation_outputs, "kpi_calculation_detail", "20_KPI_Calculation_Detail")
        _write_calculation_sheet(writer, calculation_outputs, "cpt_calculation_detail", "21_CPT_Calculation_Detail")
        _write_calculation_sheet(writer, calculation_outputs, "drg_calculation_detail", "22_DRG_Calculation_Detail")
        _write_calculation_sheet(writer, calculation_outputs, "referral_calculation_detail", "23_Referral_Calculation_Detail")
        _write_calculation_sheet(writer, calculation_outputs, "geographic_catchment_detail", "24_Geographic_Catchment_Detail")

        _format_workbook(writer.book)

    return path


def _build_read_me(account_name: str, generation_timestamp: str) -> pd.DataFrame:
    rows = [
        ["Account name", account_name],
        ["Workbook purpose", "Auditor Workbook / source fact book"],
        ["Generation timestamp", generation_timestamp],
        [
            "Statement",
            "This workbook preserves source facts and traceability. It does not calculate final KPIs, recommendations, opportunity scores, or PowerPoint outputs.",
        ],
        ["Included raw tabs", ""],
    ]
    rows.extend([[raw_tab_name, ""] for _, raw_tab_name in RAW_TAB_SHEET_ORDER])
    return pd.DataFrame(rows)


def _write_calculation_sheet(writer, calculation_outputs: dict[str, pd.DataFrame] | None, key: str, sheet_name: str) -> None:
    df = calculation_outputs.get(key, pd.DataFrame()) if calculation_outputs else pd.DataFrame()
    if df.empty:
        pd.DataFrame({"note": [f"No {sheet_name} data generated."]}).to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def _build_data_quality_log(
    source_file_inventory: pd.DataFrame,
    raw_tab_dataframes: dict[str, pd.DataFrame],
    cpt_taxonomy_matches: pd.DataFrame | None = None,
    drg_taxonomy_matches: pd.DataFrame | None = None,
    taxonomy_match_summary: pd.DataFrame | None = None,
    kpi_eligibility: pd.DataFrame | None = None,
) -> pd.DataFrame:
    loaded_count = int((source_file_inventory.get("load_status", pd.Series(dtype=str)) == "loaded").sum())
    unknown_count = int((source_file_inventory.get("detected_file_type", pd.Series(dtype=str)) == "unknown").sum())
    years = set(source_file_inventory.get("reporting_year", pd.Series(dtype=str)).dropna().astype(str))
    years.discard("")
    manual_overrides = int((source_file_inventory.get("data_year_status", pd.Series(dtype=str)) == "manual_override").sum())
    empty_tabs = [raw_tab_name for _, raw_tab_name in RAW_TAB_SHEET_ORDER if raw_tab_dataframes.get(raw_tab_name, pd.DataFrame()).empty]
    numeric_columns = sum(1 for df in raw_tab_dataframes.values() for column in df.columns if str(column).endswith("__num"))

    rows = [
            {
                "severity": "info",
                "check_name": "files_loaded",
                "status": "pass" if loaded_count > 0 else "warn",
                "detail": f"{loaded_count} file(s) loaded.",
            },
            {
                "severity": "warning",
                "check_name": "unknown_files_present",
                "status": "warn" if unknown_count > 0 else "pass",
                "detail": f"{unknown_count} unknown file(s) present.",
            },
            {
                "severity": "warning",
                "check_name": "mixed_years_present",
                "status": "warn" if len(years) > 1 else "pass",
                "detail": f"Detected year(s): {', '.join(sorted(years)) if years else 'none'}.",
            },
            {
                "severity": "Yellow" if manual_overrides else "info",
                "check_name": "filename_year_reporting_year_mismatch",
                "status": "Review" if manual_overrides else "pass",
                "detail": f"{manual_overrides} file(s) have reporting year manual overrides.",
            },
            {
                "severity": "info",
                "check_name": "empty_raw_tabs",
                "status": "warn" if empty_tabs else "pass",
                "detail": f"{len(empty_tabs)} empty raw tab(s): {', '.join(empty_tabs)}.",
            },
            {
                "severity": "info",
                "check_name": "numeric_cleaning_columns_created",
                "status": "pass" if numeric_columns > 0 else "warn",
                "detail": f"{numeric_columns} numeric companion column(s) created.",
            },
    ]
    rows.extend(_taxonomy_quality_rows(cpt_taxonomy_matches, drg_taxonomy_matches, taxonomy_match_summary))
    rows.extend(_kpi_eligibility_quality_rows(kpi_eligibility))
    return pd.DataFrame(rows, columns=["severity", "check_name", "status", "detail"])


def _taxonomy_quality_rows(
    cpt_taxonomy_matches: pd.DataFrame | None,
    drg_taxonomy_matches: pd.DataFrame | None,
    taxonomy_match_summary: pd.DataFrame | None,
) -> list[dict[str, str]]:
    cpt_df = cpt_taxonomy_matches if cpt_taxonomy_matches is not None else pd.DataFrame()
    drg_df = drg_taxonomy_matches if drg_taxonomy_matches is not None else pd.DataFrame()
    summary = taxonomy_match_summary if taxonomy_match_summary is not None else pd.DataFrame()

    cpt_matched = _status_count(cpt_df, "matched")
    cpt_unmatched = _status_count(cpt_df, "unmatched")
    drg_matched = _status_count(drg_df, "matched")
    drg_unmatched = _status_count(drg_df, "unmatched")
    cpt_rate = _overall_match_rate(summary, "CPT")
    drg_rate = _overall_match_rate(summary, "DRG")

    return [
        {"severity": "info", "check_name": "cpt_taxonomy_loaded", "status": "pass" if cpt_df is not None else "warn", "detail": f"{len(cpt_df)} CPT taxonomy row match record(s)."},
        {"severity": "info", "check_name": "drg_taxonomy_loaded", "status": "pass" if drg_df is not None else "warn", "detail": f"{len(drg_df)} DRG taxonomy row match record(s)."},
        {"severity": "info", "check_name": "cpt_rows_matched", "status": "pass", "detail": f"{cpt_matched} CPT row(s) matched."},
        {"severity": "warning", "check_name": "cpt_rows_unmatched", "status": "Review" if cpt_unmatched else "pass", "detail": f"{cpt_unmatched} CPT row(s) unmatched."},
        {"severity": "info", "check_name": "drg_rows_matched", "status": "pass", "detail": f"{drg_matched} DRG row(s) matched."},
        {"severity": "warning", "check_name": "drg_rows_unmatched", "status": "Review" if drg_unmatched else "pass", "detail": f"{drg_unmatched} DRG row(s) unmatched."},
        {"severity": "info", "check_name": "cpt_match_rate", "status": "pass", "detail": f"{cpt_rate:.4f}"},
        {"severity": "info", "check_name": "drg_match_rate", "status": "pass", "detail": f"{drg_rate:.4f}"},
    ]


def _status_count(df: pd.DataFrame, status: str) -> int:
    if df.empty or "taxonomy_match_status" not in df.columns:
        return 0
    return int((df["taxonomy_match_status"] == status).sum())


def _overall_match_rate(summary: pd.DataFrame, taxonomy_type: str) -> float:
    if summary.empty:
        return 0
    filtered = summary[summary["taxonomy_type"] == taxonomy_type]
    total = int(filtered["total_rows"].sum()) if "total_rows" in filtered else 0
    no_code = int(filtered["no_code_column_rows"].sum()) if "no_code_column_rows" in filtered else 0
    matched = int(filtered["matched_rows"].sum()) if "matched_rows" in filtered else 0
    denominator = total - no_code
    return round(matched / denominator, 4) if denominator else 0


def _kpi_eligibility_quality_rows(kpi_eligibility: pd.DataFrame | None) -> list[dict[str, str]]:
    df = kpi_eligibility if kpi_eligibility is not None else pd.DataFrame()
    eligible = _eligibility_count(df, "eligible")
    caution = _eligibility_count(df, "eligible_with_caution")
    not_eligible = _eligibility_count(df, "not_eligible")
    return [
        {"severity": "info", "check_name": "kpi_eligibility_created", "status": "pass" if not df.empty else "warn", "detail": f"{len(df)} KPI eligibility row(s) created."},
        {"severity": "info", "check_name": "kpis_eligible_count", "status": "pass", "detail": str(eligible)},
        {"severity": "info", "check_name": "kpis_caution_count", "status": "pass", "detail": str(caution)},
        {"severity": "info", "check_name": "kpis_not_eligible_count", "status": "pass", "detail": str(not_eligible)},
    ]


def _eligibility_count(df: pd.DataFrame, status: str) -> int:
    if df.empty or "eligibility_status" not in df.columns:
        return 0
    return int((df["eligibility_status"] == status).sum())


def _format_workbook(workbook) -> None:
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        if worksheet.max_row > 1 and worksheet.max_column > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_length + 2, 12), 48)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
