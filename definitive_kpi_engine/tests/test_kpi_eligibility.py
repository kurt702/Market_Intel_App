import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from definitive_engine.validation import build_kpi_eligibility
from definitive_engine.workbooks import write_auditor_workbook


def _metric(kpi_id: str) -> pd.DataFrame:
    metrics = pd.read_csv(Path("config") / "metric_definitions.csv", dtype=str, keep_default_na=False)
    return metrics[metrics["kpi_id"] == kpi_id]


class KpiEligibilityTests(unittest.TestCase):
    def test_eligible_cpt_based_kpi_when_matching_category_exists(self):
        raw_tabs = {"Raw_MedicalClaims_CPT": pd.DataFrame({"Actual Charges__num": [100.0]})}
        cpt = pd.DataFrame({"ep_category": ["Device Creation"], "taxonomy_match_status": ["matched"]})
        result = build_kpi_eligibility(_metric("device_creation_revenue"), raw_tabs, cpt, pd.DataFrame())
        self.assertEqual(result.loc[0, "eligibility_status"], "eligible")
        self.assertEqual(result.loc[0, "confidence_status"], "green")

    def test_not_eligible_cpt_based_kpi_when_category_missing(self):
        raw_tabs = {"Raw_MedicalClaims_CPT": pd.DataFrame({"Actual Charges__num": [100.0]})}
        cpt = pd.DataFrame({"ep_category": ["Device Lifecycle"], "taxonomy_match_status": ["matched"]})
        result = build_kpi_eligibility(_metric("device_creation_revenue"), raw_tabs, cpt, pd.DataFrame())
        self.assertEqual(result.loc[0, "eligibility_status"], "not_eligible")
        self.assertEqual(result.loc[0, "confidence_status"], "red")

    def test_eligible_with_caution_for_drg_based_kpi(self):
        raw_tabs = {"Raw_MedicalClaims_DRG": pd.DataFrame({"Actual Charges__num": [100.0]})}
        drg = pd.DataFrame({"drg_category": ["Core CIED Implant"], "taxonomy_match_status": ["matched"]})
        result = build_kpi_eligibility(_metric("device_implant_revenue"), raw_tabs, pd.DataFrame(), drg)
        self.assertEqual(result.loc[0, "eligibility_status"], "eligible_with_caution")
        self.assertEqual(result.loc[0, "confidence_status"], "yellow")

    def test_referral_balance_eligibility(self):
        raw_tabs = {
            "Raw_Referrals_From": pd.DataFrame({"# of referrals__num": [5]}),
            "Raw_Referrals_To": pd.DataFrame({"# of referrals__num": [7]}),
        }
        result = build_kpi_eligibility(_metric("net_referral_balance"), raw_tabs, pd.DataFrame(), pd.DataFrame())
        self.assertEqual(result.loc[0, "eligibility_status"], "eligible")

    def test_geographic_catchment_eligibility(self):
        raw_tabs = {"Raw_ZIP_Origination": pd.DataFrame({"Medicare Total # of Cases__num": [10]})}
        result = build_kpi_eligibility(_metric("geographic_catchment"), raw_tabs, pd.DataFrame(), pd.DataFrame())
        self.assertEqual(result.loc[0, "eligibility_status"], "eligible")

    def test_kpi_eligibility_sheet_written_to_auditor_workbook(self):
        with TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "auditor_workbook.xlsx"
            kpi_eligibility = pd.DataFrame(
                [
                    {
                        "kpi_id": "device_creation_revenue",
                        "kpi_name": "Device Creation Revenue",
                        "domain": "Device Creation",
                        "required_source_tabs": "Raw_MedicalClaims_CPT",
                        "required_taxonomy_type": "CPT",
                        "required_code_category": "Device Creation",
                        "required_numeric_fields": "Actual Charges__num",
                        "source_tabs_present": True,
                        "required_fields_present": True,
                        "taxonomy_matches_present": True,
                        "eligibility_status": "eligible",
                        "confidence_status": "green",
                        "eligibility_notes": "Test row.",
                    }
                ]
            )
            write_auditor_workbook(
                output,
                "Rush",
                pd.DataFrame(),
                {},
                pd.DataFrame(),
                kpi_eligibility=kpi_eligibility,
            )
            workbook = load_workbook(output, read_only=True)
            self.assertIn("19_KPI_Eligibility", workbook.sheetnames)
            self.assertEqual(workbook["19_KPI_Eligibility"]["A2"].value, "device_creation_revenue")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
