import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from definitive_engine.calculations import build_calculation_outputs
from definitive_engine.validation import build_kpi_eligibility
from definitive_engine.workbooks import write_auditor_workbook


def _eligibility(ids):
    metrics = pd.read_csv(Path("config") / "metric_definitions.csv", dtype=str, keep_default_na=False)
    raw_tabs = {
        "Raw_MedicalClaims_CPT": pd.DataFrame({"Actual Charges__num": [1.0]}),
        "Raw_MedicalClaims_DRG": pd.DataFrame({"Actual Charges__num": [1.0], "# Actual Claims__num": [1.0]}),
        "Raw_Inpatient_Leakage": pd.DataFrame({"Total Charges__num": [1.0]}),
        "Raw_Referrals_From": pd.DataFrame({"# of referrals__num": [1.0]}),
        "Raw_Referrals_To": pd.DataFrame({"# of referrals__num": [1.0]}),
        "Raw_ZIP_Origination": pd.DataFrame({"Medicare Total # of Cases__num": [1.0]}),
        "Raw_Inpatient_MarketShare": pd.DataFrame({"% claims at this hospital__num": [1.0], "Total claims - patient universe__num": [1.0]}),
    }
    cpt = pd.DataFrame({"ep_category": ["Device Creation", "Device Lifecycle", "Lead Management"], "taxonomy_match_status": ["matched", "matched", "matched"]})
    drg = pd.DataFrame({"drg_category": ["Core CIED Implant"], "taxonomy_match_status": ["matched"], "source_tab": ["Raw_MedicalClaims_DRG"]})
    elig = build_kpi_eligibility(metrics[metrics["kpi_id"].isin(ids)], raw_tabs, cpt, drg)
    return elig


def _base_frames():
    cpt = pd.DataFrame(
        {
            "source_file": ["cpt.csv", "cpt.csv", "cpt.csv"],
            "source_row_number": [2, 3, 4],
            "source_tab": ["Raw_MedicalClaims_CPT"] * 3,
            "HCPCS/CPT Code": ["33249", "33228", "33244"],
            "HCPCS/CPT Description": ["Create", "Lifecycle", "Lead"],
            "cpt_code_normalized": ["33249", "33228", "33244"],
            "ep_category": ["Device Creation", "Device Lifecycle", "Lead Management"],
            "kpi_domain": ["Device Creation", "Device Lifecycle", "Lead Management"],
            "include_in_total_ep_charges": ["TRUE", "TRUE", "TRUE"],
            "taxonomy_match_status": ["matched", "matched", "matched"],
            "Actual Charges__num": [100.0, 200.0, 50.0],
            "# Actual Procedures__num": [1.0, 2.0, 1.0],
        }
    )
    drg = pd.DataFrame(
        {
            "source_tab": ["Raw_MedicalClaims_DRG", "Raw_Inpatient_Leakage", "Raw_Inpatient_MarketShare"],
            "source_file": ["drg.csv", "leak.csv", "share.csv"],
            "source_row_number": [2, 2, 2],
            "DRG Code": ["242", "", "242"],
            "Base DRG Group": ["", "242 - Permanent cardiac pacemaker implant", ""],
            "drg_code_normalized": ["242", "242", "242"],
            "drg_category": ["Core CIED Implant", "Core CIED Implant", "Core CIED Implant"],
            "kpi_domain": ["Device Implant Revenue", "Leakage", "Market Position"],
            "recommended_use": ["include", "include", "include"],
            "taxonomy_match_status": ["matched", "matched", "matched"],
            "Actual Charges__num": [1000.0, None, None],
            "# Actual Claims__num": [10.0, None, None],
            "Total Charges__num": [None, 500.0, None],
            "% claims at this hospital__num": [None, None, 25.0],
            "Total claims - this hospital__num": [None, None, 25.0],
            "Total claims - patient universe__num": [None, None, 100.0],
        }
    )
    raw_tabs = {
        "Raw_MedicalClaims_CPT": pd.DataFrame({"Actual Charges__num": [1]}),
        "Raw_MedicalClaims_DRG": pd.DataFrame({"Actual Charges__num": [1], "# Actual Claims__num": [1]}),
        "Raw_Inpatient_Leakage": pd.DataFrame({"Total Charges__num": [1]}),
        "Raw_Referrals_From": pd.DataFrame({"# of referrals__num": [10, 5], "source_file": ["from.csv", "from.csv"]}),
        "Raw_Referrals_To": pd.DataFrame({"# of referrals__num": [7], "source_file": ["to.csv"]}),
        "Raw_ZIP_Origination": pd.DataFrame({"Zip Code": [str(i) for i in range(30)], "City": ["C"] * 30, "State": ["IL"] * 30, "Medicare Total # of Cases__num": list(range(30)), "Medicare Total Charges__num": list(range(100, 130))}),
        "Raw_Inpatient_MarketShare": pd.DataFrame({"% claims at this hospital__num": [25], "Total claims - patient universe__num": [100]}),
    }
    return raw_tabs, cpt, drg


class CalculationTests(unittest.TestCase):
    def _detail(self, ids):
        raw_tabs, cpt, drg = _base_frames()
        return build_calculation_outputs(_eligibility(ids), raw_tabs, cpt, drg)

    def test_cpt_revenue_calculations(self):
        detail = self._detail(["device_creation_revenue", "device_lifecycle_revenue", "lead_management_revenue"])["kpi_calculation_detail"]
        values = dict(zip(detail["kpi_id"], detail["calculated_value"]))
        self.assertEqual(values["device_creation_revenue"], 100.0)
        self.assertEqual(values["device_lifecycle_revenue"], 200.0)
        self.assertEqual(values["lead_management_revenue"], 50.0)

    def test_total_ep_charges_and_dpci(self):
        detail = self._detail(["total_ep_charges", "dpci"])["kpi_calculation_detail"]
        values = dict(zip(detail["kpi_id"], detail["calculated_value"]))
        self.assertEqual(values["total_ep_charges"], 350.0)
        self.assertEqual(values["dpci"], 300.0)
        total_ep = detail[detail["kpi_id"] == "total_ep_charges"].iloc[0]
        self.assertEqual(total_ep["kpi_name"], "Mapped EP/CIED Charges")
        self.assertIn("mapped EP/CIED CPT taxonomy only", total_ep["validation_notes"])

    def test_device_implant_revenue_and_volume(self):
        detail = self._detail(["device_implant_revenue", "device_implant_volume"])["kpi_calculation_detail"]
        values = dict(zip(detail["kpi_id"], detail["calculated_value"]))
        self.assertEqual(values["device_implant_revenue"], 1000.0)
        self.assertEqual(values["device_implant_volume"], 10.0)

    def test_ep_leakage_revenue(self):
        detail = self._detail(["ep_leakage_revenue"])["kpi_calculation_detail"]
        self.assertEqual(detail.loc[0, "calculated_value"], 500.0)

    def test_net_referral_balance(self):
        detail = self._detail(["net_referral_balance"])
        row = detail["kpi_calculation_detail"].iloc[0]
        self.assertEqual(row["calculated_value"], 8.0)
        self.assertEqual(len(detail["referral_calculation_detail"]), 3)

    def test_geographic_catchment_top_25(self):
        detail = self._detail(["geographic_catchment"])
        self.assertEqual(len(detail["geographic_catchment_detail"]), 25)
        self.assertEqual(detail["kpi_calculation_detail"].iloc[0]["calculated_value"], sum(range(5, 30)))

    def test_market_rank_not_calculated(self):
        detail = self._detail(["market_rank"])["kpi_calculation_detail"]
        self.assertEqual(detail.loc[0, "calculation_status"], "not_calculated")

    def test_ep_market_share_uses_numerator_denominator(self):
        detail = self._detail(["ep_market_share"])["kpi_calculation_detail"]
        row = detail.iloc[0]
        self.assertEqual(row["calculation_status"], "calculated_with_caution")
        self.assertEqual(row["calculated_value"], 25.0)
        self.assertIn("claims numerator/denominator", row["validation_notes"])

    def test_ep_market_share_not_calculated_from_percentage_alone(self):
        raw_tabs, cpt, drg = _base_frames()
        drg = drg.drop(columns=["Total claims - this hospital__num", "Total claims - patient universe__num"], errors="ignore")
        eligibility = _eligibility(["ep_market_share"])
        outputs = build_calculation_outputs(eligibility, raw_tabs, cpt, drg)
        row = outputs["kpi_calculation_detail"].iloc[0]
        self.assertEqual(row["calculation_status"], "not_calculated")
        self.assertIn("not calculated from percentage fields alone", row["validation_notes"])

    def test_calculation_detail_sheets_written(self):
        with TemporaryDirectory() as temp_dir:
            outputs = self._detail(["device_creation_revenue"])
            path = Path(temp_dir) / "auditor_workbook.xlsx"
            write_auditor_workbook(path, "Rush", pd.DataFrame(), {}, pd.DataFrame(), calculation_outputs=outputs)
            wb = load_workbook(path, read_only=True)
            self.assertIn("20_KPI_Calculation_Detail", wb.sheetnames)
            self.assertIn("21_CPT_Calculation_Detail", wb.sheetnames)
            self.assertIn("22_DRG_Calculation_Detail", wb.sheetnames)
            self.assertIn("23_Referral_Calculation_Detail", wb.sheetnames)
            self.assertIn("24_Geographic_Catchment_Detail", wb.sheetnames)
            wb.close()


if __name__ == "__main__":
    unittest.main()
