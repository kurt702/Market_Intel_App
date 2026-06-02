import json
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from definitive_engine.final_outputs import KPI_WORKBOOK_SHEETS, write_final_outputs


class FinalOutputsTests(unittest.TestCase):
    def test_final_outputs_created_with_expected_content(self):
        with TemporaryDirectory() as temp_dir:
            outputs = _calculation_outputs()
            paths = write_final_outputs(temp_dir, "Rush", "2024", outputs, "auditor_workbook.xlsx")

            self.assertTrue(paths["kpi_workbook"].exists())
            self.assertTrue(paths["kpi_summary_csv"].exists())
            self.assertTrue(paths["kpi_summary_json"].exists())

            wb = load_workbook(paths["kpi_workbook"], read_only=True)
            self.assertEqual(wb.sheetnames, KPI_WORKBOOK_SHEETS)
            wb.close()

            csv = pd.read_csv(paths["kpi_summary_csv"])
            self.assertEqual(len(csv), 12)
            self.assertIn("Mapped EP/CIED Charges", set(csv["kpi_name"]))

            payload = json.loads(paths["kpi_summary_json"].read_text(encoding="utf-8"))
            self.assertEqual(len(payload["kpis"]), 12)

            market_rank = csv[csv["kpi_id"] == "market_rank"].iloc[0]
            self.assertEqual(market_rank["formatted_value"], "Not calculated")
            self.assertEqual(market_rank["calculation_status"], "not_calculated")

            self.assertFalse(any(Path(temp_dir).glob("*.pptx")))
            self.assertFalse(any(Path(temp_dir).glob("*.ppt")))


def _calculation_outputs():
    detail = pd.DataFrame(
        [
            _row("total_ep_charges", "Mapped EP/CIED Charges", "EP Financial", 1000000, "dollars", "eligible_with_caution", "calculated_with_caution", "yellow"),
            _row("device_implant_revenue", "Device Implant Revenue", "Device Implant Revenue", 2000000, "dollars", "eligible_with_caution", "calculated_with_caution", "yellow"),
            _row("device_implant_volume", "Device Implant Volume", "Device Implant Volume", 20, "cases/claims", "eligible_with_caution", "calculated_with_caution", "yellow"),
            _row("device_creation_revenue", "Device Creation Revenue", "Device Creation", 300, "dollars", "eligible", "calculated", "green"),
            _row("device_lifecycle_revenue", "Device Lifecycle Revenue", "Device Lifecycle", 400, "dollars", "eligible", "calculated", "green"),
            _row("lead_management_revenue", "Lead Management Revenue", "Lead Management", 50, "dollars", "eligible", "calculated", "green"),
            _row("ep_leakage_revenue", "EP Leakage Revenue", "Leakage", 5000, "dollars", "eligible_with_caution", "calculated_with_caution", "yellow"),
            _row("net_referral_balance", "Net Referral Balance", "Referrals", 3, "referrals", "eligible", "calculated", "green"),
            _row("geographic_catchment", "Geographic Catchment", "Geography", 25, "cases", "eligible", "calculated", "green"),
            _row("ep_market_share", "EP Market Share", "Market Position", 44.3, "percent", "eligible_with_caution", "calculated_with_caution", "yellow"),
            _row("market_rank", "Market Rank", "Market Position", "", "rank", "eligible_with_caution", "not_calculated", "yellow"),
            _row("dpci", "DPCI", "Data Completeness", 700, "dollars", "eligible", "calculated", "green"),
        ]
    )
    return {
        "kpi_calculation_detail": detail,
        "cpt_calculation_detail": pd.DataFrame({"cpt_code_normalized": ["33249"], "ep_category": ["Device Creation"], "kpi_domain": ["Device Creation"], "Actual Charges__num": [300], "# Actual Procedures__num": [1]}),
        "drg_calculation_detail": pd.DataFrame({"source_tab": ["Raw_MedicalClaims_DRG"], "drg_code_normalized": ["242"], "drg_category": ["Core CIED Implant"], "kpi_domain": ["Device Implant Revenue"], "recommended_use": ["include"], "Actual Charges__num": [2000000], "# Actual Claims__num": [20]}),
        "referral_calculation_detail": pd.DataFrame({"component": ["incoming_referrals", "outgoing_referrals", "net_balance"], "referrals": [10, 7, 3], "source_files": ["to.csv", "from.csv", "from.csv; to.csv"]}),
        "geographic_catchment_detail": pd.DataFrame({"ZIP Code": ["60612"], "City": ["Chicago"], "State": ["IL"], "Medicare Total # of Cases__num": [25], "Medicare Total Charges__num": [1000], "rank": [1]}),
    }


def _row(kpi_id, name, domain, value, unit, eligibility, calc, confidence):
    return {
        "kpi_id": kpi_id,
        "kpi_name": name,
        "domain": domain,
        "eligibility_status": eligibility,
        "confidence_status": confidence,
        "calculation_status": calc,
        "calculated_value": value,
        "unit": unit,
        "source_tab": "source",
        "source_file": "file.csv",
        "source_row_count": 1,
        "matched_row_count": 1,
        "calculation_method": "method",
        "source_value_field": "field",
        "source_count_field": "",
        "taxonomy_filter": "filter",
        "validation_notes": "note",
    }


if __name__ == "__main__":
    unittest.main()
