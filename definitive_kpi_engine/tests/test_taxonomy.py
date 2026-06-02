import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from definitive_engine.ingestion import build_raw_tabs
from definitive_engine.raw_tabs import build_raw_tab_summary
from definitive_engine.taxonomy import (
    apply_cpt_taxonomy,
    apply_drg_taxonomy,
    build_taxonomy_outputs,
    normalize_cpt_code,
    normalize_drg_code,
)
from definitive_engine.workbooks import write_auditor_workbook


CPT_TAXONOMY = pd.DataFrame(
    {
        "cpt_hcpcs_code": ["33249"],
        "ep_category": ["EP Device"],
        "kpi_domain": ["EP"],
        "include_in_total_ep_charges": ["true"],
        "include_in_dpci": ["true"],
    }
)

DRG_TAXONOMY = pd.DataFrame(
    {
        "drg_code": ["242"],
        "drg_category": ["Pacemaker"],
        "kpi_domain": ["EP"],
        "recommended_use": ["include"],
    }
)


class TaxonomyTests(unittest.TestCase):
    def test_cpt_code_normalization(self):
        self.assertEqual(normalize_cpt_code("33249"), "33249")
        self.assertEqual(normalize_cpt_code("33249.0"), "33249")
        self.assertEqual(normalize_cpt_code(" 33249 "), "33249")

    def test_drg_code_normalization(self):
        self.assertEqual(normalize_drg_code("242"), "242")
        self.assertEqual(normalize_drg_code("242.0"), "242")
        self.assertEqual(normalize_drg_code("242 - Permanent cardiac pacemaker implant"), "242")

    def test_cpt_taxonomy_matching_and_unmatched_rows_preserved(self):
        df = pd.DataFrame({"HCPCS/CPT Code": ["33249", "99999"], "source_file": ["a.csv", "a.csv"], "source_row_number": [2, 3]})
        result = apply_cpt_taxonomy(df, CPT_TAXONOMY, source_tab="Raw_MedicalClaims_CPT")
        self.assertEqual(len(result), 2)
        self.assertEqual(result.loc[0, "taxonomy_match_status"], "matched")
        self.assertEqual(result.loc[0, "ep_category"], "EP Device")
        self.assertEqual(result.loc[1, "taxonomy_match_status"], "unmatched")

    def test_drg_taxonomy_matching_and_unmatched_rows_preserved(self):
        df = pd.DataFrame({"DRG Code": ["242", "999"], "source_file": ["a.csv", "a.csv"], "source_row_number": [2, 3]})
        result = apply_drg_taxonomy(df, DRG_TAXONOMY, source_tab="Raw_MedicalClaims_DRG")
        self.assertEqual(len(result), 2)
        self.assertEqual(result.loc[0, "taxonomy_match_status"], "matched")
        self.assertEqual(result.loc[0, "drg_category"], "Pacemaker")
        self.assertEqual(result.loc[1, "taxonomy_match_status"], "unmatched")

    def test_no_cpt_column_behavior(self):
        df = pd.DataFrame({"Procedure": ["No code"], "source_file": ["a.csv"], "source_row_number": [2]})
        result = apply_cpt_taxonomy(df, CPT_TAXONOMY, source_tab="Raw_MedicalClaims_CPT")
        self.assertEqual(result.loc[0, "taxonomy_match_status"], "no_cpt_column")

    def test_no_drg_column_behavior(self):
        df = pd.DataFrame({"Diagnosis": ["No code"], "source_file": ["a.csv"], "source_row_number": [2]})
        result = apply_drg_taxonomy(df, DRG_TAXONOMY, source_tab="Raw_MedicalClaims_DRG")
        self.assertEqual(result.loc[0, "taxonomy_match_status"], "no_drg_column")

    def test_taxonomy_sheets_are_written_to_auditor_workbook(self):
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            (input_dir / "Rush_CPT_2025.csv").write_text(
                "Rendering Provider,Definitive ID,Rendering Provider Type,HCPCS/CPT Code,HCPCS/CPT Description,# Actual Procedures,% Actual Procedures,Actual Charges,% Actual Charges,Charge/Procedure (Actual)\n"
                "Dr Example,123,Physician,33249,Device,20,5,10000,4,500\n",
                encoding="utf-8",
            )
            inventory, raw_tabs = build_raw_tabs("Rush", input_dir, reporting_year="2024")
            cpt_matches, drg_matches, taxonomy_summary = build_taxonomy_outputs(raw_tabs, CPT_TAXONOMY, DRG_TAXONOMY)
            workbook_path = write_auditor_workbook(
                output_dir / "auditor_workbook.xlsx",
                "Rush",
                inventory,
                raw_tabs,
                build_raw_tab_summary(raw_tabs),
                cpt_taxonomy_matches=cpt_matches,
                drg_taxonomy_matches=drg_matches,
                taxonomy_match_summary=taxonomy_summary,
            )
            workbook = load_workbook(workbook_path, read_only=True)
            self.assertIn("16_CPT_Taxonomy_Matches", workbook.sheetnames)
            self.assertIn("17_DRG_Taxonomy_Matches", workbook.sheetnames)
            self.assertEqual(workbook["16_CPT_Taxonomy_Matches"]["G2"].value, "33249")
            workbook.close()

            created_names = {path.name for path in output_dir.iterdir()}
            self.assertNotIn("kpi_workbook.xlsx", created_names)
            self.assertNotIn("kpi_summary.csv", created_names)
            self.assertNotIn("kpi_summary.json", created_names)
            self.assertFalse(any(path.suffix.lower() in {".ppt", ".pptx"} for path in output_dir.iterdir()))


if __name__ == "__main__":
    unittest.main()

