import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from definitive_engine.ingestion import build_raw_tabs
from definitive_engine.raw_tabs import build_raw_tab_summary
from definitive_engine.workbooks import AUDITOR_SHEET_NAMES, write_auditor_workbook


MEDICALCLAIMS_CPT_HEADER = [
    "Rendering Provider",
    "Definitive ID",
    "Rendering Provider Type",
    "HCPCS/CPT Code",
    "HCPCS/CPT Description",
    "# Actual Procedures",
    "% Actual Procedures",
    "Actual Charges",
    "% Actual Charges",
    "Charge/Procedure (Actual)",
]


class AuditorWorkbookTests(unittest.TestCase):
    def test_workbook_file_and_expected_sheets_are_created(self):
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            (input_dir / "Rush_CPT_2025.csv").write_text(
                ",".join(MEDICALCLAIMS_CPT_HEADER) + "\n" + ",".join(["value"] * len(MEDICALCLAIMS_CPT_HEADER)) + "\n",
                encoding="utf-8",
            )
            inventory, raw_tabs = build_raw_tabs("Rush", input_dir, load_timestamp="2026-06-02T00:00:00+00:00")
            raw_tab_summary = build_raw_tab_summary(raw_tabs)

            workbook_path = write_auditor_workbook(output_dir / "auditor_workbook.xlsx", "Rush", inventory, raw_tabs, raw_tab_summary)

            self.assertTrue(workbook_path.exists())
            workbook = load_workbook(workbook_path, read_only=True)
            self.assertEqual(workbook.sheetnames, AUDITOR_SHEET_NAMES)
            self.assertIn("02_Source_File_Inventory", workbook.sheetnames)
            self.assertIn("05_Raw_MedicalClaims_CPT", workbook.sheetnames)
            self.assertIn("15_Raw_Unknown", workbook.sheetnames)
            workbook.close()

    def test_empty_raw_tabs_are_created_with_note(self):
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output_dir = root / "output"
            workbook_path = write_auditor_workbook(
                output_dir / "auditor_workbook.xlsx",
                "Rush",
                source_file_inventory=_empty_inventory(),
                raw_tab_dataframes={},
                raw_tab_summary=_empty_raw_tab_summary(),
            )

            workbook = load_workbook(workbook_path, read_only=True)
            sheet = workbook["05_Raw_MedicalClaims_CPT"]
            self.assertEqual(sheet["A1"].value, "note")
            self.assertEqual(sheet["A2"].value, "No source data loaded for this tab.")
            workbook.close()

    def test_no_kpi_or_powerpoint_outputs_are_created(self):
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output_dir = root / "output"
            write_auditor_workbook(
                output_dir / "auditor_workbook.xlsx",
                "Rush",
                source_file_inventory=_empty_inventory(),
                raw_tab_dataframes={},
                raw_tab_summary=_empty_raw_tab_summary(),
            )

            created_names = {path.name for path in output_dir.iterdir()}
            self.assertIn("auditor_workbook.xlsx", created_names)
            self.assertNotIn("kpi_workbook.xlsx", created_names)
            self.assertNotIn("kpi_summary.csv", created_names)
            self.assertNotIn("kpi_summary.json", created_names)
            self.assertFalse(any(path.suffix.lower() in {".ppt", ".pptx"} for path in output_dir.iterdir()))

    def test_data_quality_log_flags_reporting_year_manual_override(self):
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            (input_dir / "Rush_CPT_2025.csv").write_text(
                ",".join(MEDICALCLAIMS_CPT_HEADER) + "\n" + ",".join(["value"] * len(MEDICALCLAIMS_CPT_HEADER)) + "\n",
                encoding="utf-8",
            )
            inventory, raw_tabs = build_raw_tabs(
                "Rush",
                input_dir,
                load_timestamp="2026-06-02T00:00:00+00:00",
                reporting_year="2024",
            )
            raw_tab_summary = build_raw_tab_summary(raw_tabs)
            workbook_path = write_auditor_workbook(output_dir / "auditor_workbook.xlsx", "Rush", inventory, raw_tabs, raw_tab_summary)

            workbook = load_workbook(workbook_path, read_only=True)
            rows = list(workbook["04_Data_Quality_Log"].iter_rows(values_only=True))
            header = rows[0]
            log_rows = [dict(zip(header, row)) for row in rows[1:]]
            mismatch = next(row for row in log_rows if row["check_name"] == "filename_year_reporting_year_mismatch")
            self.assertEqual(mismatch["severity"], "Yellow")
            self.assertEqual(mismatch["status"], "Review")
            workbook.close()


def _empty_inventory():
    import pandas as pd

    return pd.DataFrame(
        columns=[
            "account_name",
            "file_name",
            "file_path",
            "row_count",
            "column_count",
            "column_names",
            "filename_year",
            "reporting_year",
            "data_year_status",
            "data_year_notes",
            "load_status",
            "detected_file_type",
            "classification_notes",
        ]
    )


def _empty_raw_tab_summary():
    import pandas as pd

    return pd.DataFrame(columns=["raw_tab_name", "detected_file_type", "row_count", "column_count", "source_files"])


if __name__ == "__main__":
    unittest.main()
